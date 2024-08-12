import re
import sys
from PyQt5 import QtWidgets
from PyQt5.QtCore import QTimer, lowercasedigits
from PyQt5.QtGui import QFont
from designQT import Ui_Form
import time
import datetime
import nfc
from playsound import playsound
clf = nfc.ContactlessFrontend('usb')
from nfc.clf import RemoteTarget
import openpyxl
import threading
import designQT
from goto import with_goto
import pdfkit, os
import platform
from smb.SMBConnection import SMBConnection

# wkhtmltopdfのインストールパスを指定
wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'

#PDFファイル生成
def makePDF(d):
    # ファイルのパスを指定
    temp_file =  'html/'+str(d)+'.html'
    output_file = 'pdf/'+str(d)+'.pdf'
    f = open('html/'+d+".html","w",encoding="utf-8")
    #HTML本文
    datalist = '<!doctype html><html lang="jp"><head><!-- Required meta tags --><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><!-- Bootstrap CSS --></script><title>'+d+'</title></head><body><h1>'+d+'</h1><div class="container"><table class="table"><thead><tr><th scope="col">日付</th><th scope="col">学籍番号</th><th scope="col">クラス</th><th scope="col">番号</th><th scope="col">氏名</th></tr></thead><tbody>'
    #Excelを読み込み
    wb = openpyxl.load_workbook('Book1.xlsx')
    #Excelのシート読み込み
    ws = wb[d_now]

    #出席番号順にソート
    temp_number = []
    list_date = []
    list_id = []
    list_class = []
    list_number = []
    list_name = []

    new_date = []
    new_id = []
    new_class = []
    new_number = []
    new_name = []

    for row in ws.iter_rows(min_row=1):
        data_date = row[0].value
        list_date.append(data_date)
        data_id = int(row[1].value)
        list_id.append(data_id)
        data_class = row[2].value
        list_class.append(data_class)
        data_number = int(row[3].value)
        list_number.append(data_number)
        temp_number.append(data_number)
        data_name = row[4].value
        list_name.append(data_name)

    temp_number.sort()
    print(temp_number)

    for row in range(len(temp_number)):
        for i in range(len(list_number)):
            if list_number[i] == temp_number[row]:
                new_date.append(list_date[i])
                new_id.append(list_id[i])
                new_class.append(list_class[i])
                new_number.append(list_number[i])
                new_name.append(list_name[i])
                break
        
    
    #HTMLテーブル作成
    for i in range(len(new_id)):
        datalist += '<tr><th scope="row">'+str(new_date[i])+'</th><td>'+str(new_id[i])+'</td><td>'+str(new_class[i])+'</td><td>'+str(new_number[i])+'</td><td>'+str(new_name[i])+'</td></tr>'
    
    datalist += ' </tbody></table></div></body></html>'
    #HTML書き出し
    f.writelines(datalist)
    #HTML閉じる
    f.close()

    # pdf 変換オプションなどを指定
    conf = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf)
    options = {'page-size': 'A4', 'encoding': "UTF-8"}

    # HTMLファイルをPDF出力
    pdfkit.from_file(temp_file, output_file, options=options, configuration=conf)
    print("pdf-ok")

#HTML作成
def makeHTML(d):
    # ファイルのパスを指定
    f = open('_html/'+d+".html","w",encoding="utf-8")
    #HTML本文
    datalist = '<!DOCTYPE html><html lang="jp"><head><meta charset="UTF-8" /><meta name="viewport" content="width=device-width, initial-scale=1.0" /><title>'+d+'</title> <script type="text/javascript" src="./jquery-3.5.1.min.js"></script> <script type="text/javascript" src="./jquery.tablesorter.min.js"></script> <link href="./bootstrap/css/bootstrap.min.css" rel="stylesheet" /> <script src="./bootstrap/js/bootstrap.min.js"></script> <script type="text/javascript" src="./jquery-3.5.1.min.js"></script> <script type="text/javascript" src="./jquery.tablesorter.min.js"></script> <link rel="stylesheet" href="./bootstrap_table_min.css" /> <style> myTable .tablesorter-header { cursor: pointer; outline: none; } #myTable .tablesorter-header-inner::after { content: "▼"; font-size: 12px; margin-left: 5px; } .tablesorter-bootstrap { width: 100%; } .tablesorter-bootstrap tfoot td, .tablesorter-bootstrap tfoot th, .tablesorter-bootstrap thead td, .tablesorter-bootstrap thead th { font: 14px/20px Arial, Sans-serif; font-weight: 700; padding: 4px; margin: 0 0 18px; background-color: #eee; } .tablesorter-bootstrap .tablesorter-header { cursor: pointer; } .tablesorter-bootstrap .sorter-false { cursor: default; } .tablesorter-bootstrap .tablesorter-header.sorter-false i.tablesorter-icon { display: none; } .tablesorter-bootstrap .tablesorter-header-inner { position: relative; padding: 4px 18px 4px 4px; } .tablesorter-bootstrap .sorter-false .tablesorter-header-inner { padding: 4px; } .tablesorter-bootstrap .tablesorter-header i.tablesorter-icon { font-size: 11px; position: absolute; right: 2px; top: 50%; margin-top: -7px; width: 14px; height: 14px; background-repeat: no-repeat; line-height: 14px; display: inline-block; } .tablesorter-bootstrap .bootstrap-icon-unsorted { background-image: url(data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAsAAAAOCAYAAAD5YeaVAAAA20lEQVR4AWJABpKSkoxALCstLb0aUAsZaCAMhVEY6B0amx8YZWDDEDSBa2AGe7XeIiAAClYwVGBvsAcIllsf/mvcC9DgOOd8h90fxWvngVEUbZIkuWRZZlE8eQjcisgZMM9zi+LJ6ZfwegmWZflZDugdHMfxTcGqql7TNBlUB/QObtv2VBSFrev6OY7jngzFk9OT/fn73fWYpqnlXNyXDMWT0zuYx/Bvel9ej+LJ6R08DMOu67q7DkTkrSA5vYPneV71fX/QASdTkJwezhs0TfMARn0wMDDGXEPgF4oijqwM5YjNAAAAAElFTkSuQmCC); } .tablesorter-bootstrap .bootstrap-icon-white.bootstrap-icon-unsorted { background-image: url(data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAsAAAAOCAYAAAD5YeaVAAAAe0lEQVR4AbXQoRWDMBiF0Sh2QLAAQ8SxJGugWSA6A2STW1PxTsnB9cnkfuYvv8OGC1t5G3Y0QMP+Bm857keAdQIzWBP3+Bw4MADQE18B6/etRnCV/w9nnGuLezfAmXhABGtAGIkruvk6auIFRwQJDywllsEAjCecB20GP59BQQ+gtlRLAAAAAElFTkSuQmCC); } .tablesorter-bootstrap > tbody > tr.odd > td, .tablesorter-bootstrap > tbody > tr.tablesorter-hasChildRow.odd:hover ~ tr.tablesorter-hasChildRow.odd ~ .tablesorter-childRow.odd > td { background-color: #f9f9f9; } .tablesorter-bootstrap > tbody > tr.even:hover > td, .tablesorter-bootstrap > tbody > tr.hover > td, .tablesorter-bootstrap > tbody > tr.odd:hover > td, .tablesorter-bootstrap > tbody > tr.tablesorter-hasChildRow.even:hover ~ .tablesorter-childRow.even > td, .tablesorter-bootstrap > tbody > tr.tablesorter-hasChildRow.odd:hover ~ .tablesorter-childRow.odd > td { background-color: #f5f5f5; } .tablesorter-bootstrap > tbody > tr.even > td, .tablesorter-bootstrap > tbody > tr.tablesorter-hasChildRow.even:hover ~ tr.tablesorter-hasChildRow.even ~ .tablesorter-childRow.even > td { background-color: #fff; } .tablesorter-bootstrap .tablesorter-processing { background-image: url(data:image/gif;base64,R0lGODlhFAAUAKEAAO7u7lpaWgAAAAAAACH/C05FVFNDQVBFMi4wAwEAAAAh+QQBCgACACwAAAAAFAAUAAACQZRvoIDtu1wLQUAlqKTVxqwhXIiBnDg6Y4eyx4lKW5XK7wrLeK3vbq8J2W4T4e1nMhpWrZCTt3xKZ8kgsggdJmUFACH5BAEKAAIALAcAAAALAAcAAAIUVB6ii7jajgCAuUmtovxtXnmdUAAAIfkEAQoAAgAsDQACAAcACwAAAhRUIpmHy/3gUVQAQO9NetuugCFWAAAh+QQBCgACACwNAAcABwALAAACE5QVcZjKbVo6ck2AF95m5/6BSwEAIfkEAQoAAgAsBwANAAsABwAAAhOUH3kr6QaAcSrGWe1VQl+mMUIBACH5BAEKAAIALAIADQALAAcAAAIUlICmh7ncTAgqijkruDiv7n2YUAAAIfkEAQoAAgAsAAAHAAcACwAAAhQUIGmHyedehIoqFXLKfPOAaZdWAAAh+QQFCgACACwAAAIABwALAAACFJQFcJiXb15zLYRl7cla8OtlGGgUADs=); background-position: center center !important; background-repeat: no-repeat !important; } .tablesorter-bootstrap > tbody > tr.odd td.primary { background-color: #bfbfbf; } .tablesorter-bootstrap > tbody > tr td.primary, .tablesorter-bootstrap > tbody > tr.even td.primary { background-color: #d9d9d9; } .tablesorter-bootstrap > tbody > tr.odd td.secondary { background-color: #d9d9d9; } .tablesorter-bootstrap > tbody > tr td.secondary, .tablesorter-bootstrap > tbody > tr.even td.secondary { background-color: #e6e6e6; } .tablesorter-bootstrap > tbody > tr.odd td.tertiary { background-color: #e6e6e6; } .tablesorter-bootstrap > tbody > tr td.tertiary, .tablesorter-bootstrap > tbody > tr.even td.tertiary { background-color: #f2f2f2; } .tablesorter-bootstrap > .caption { background-color: #fff; } .tablesorter-bootstrap .tablesorter-filter-row input.tablesorter-filter, .tablesorter-bootstrap .tablesorter-filter-row select.tablesorter-filter { width: 98%; margin: 0; padding: 4px 6px; color: #333; -webkit-box-sizing: border-box; -moz-box-sizing: border-box; box-sizing: border-box; -webkit-transition: height 0.1s ease; -moz-transition: height 0.1s ease; -o-transition: height 0.1s ease; transition: height 0.1s ease; } .tablesorter-bootstrap .tablesorter-filter-row .tablesorter-filter.disabled { background-color: #eee; color: #555; cursor: not-allowed; border: 1px solid #ccc; border-radius: 4px; box-shadow: 0 1px 1px rgba(0, 0, 0, 0.075) inset; box-sizing: border-box; transition: height 0.1s ease; } .tablesorter-bootstrap .tablesorter-filter-row { background-color: #efefef; } .tablesorter-bootstrap .tablesorter-filter-row td { background-color: #efefef; line-height: normal; text-align: center; padding: 4px 6px; vertical-align: middle; -webkit-transition: line-height 0.1s ease; -moz-transition: line-height 0.1s ease; -o-transition: line-height 0.1s ease; transition: line-height 0.1s ease; } .tablesorter-bootstrap .tablesorter-filter-row.hideme td { padding: 2px; margin: 0; line-height: 0; } .tablesorter-bootstrap .tablesorter-filter-row.hideme * { height: 1px; min-height: 0; border: 0; padding: 0; margin: 0; opacity: 0; } .tablesorter .filtered { display: none; } .tablesorter-bootstrap .tablesorter-pager select { padding: 4px 6px; } .tablesorter-bootstrap .tablesorter-pager .pagedisplay { border: 0; } .tablesorter-bootstrap tfoot i { font-size: 11px; } .tablesorter .tablesorter-errorRow td { text-align: center; cursor: pointer; background-color: #e6bf99; } </style> <script> $(document).ready(function () { $("#myTable").tablesorter(); }); </script> </head> <body> <div class=" d-flex flex-column flex-md-row align-items-center p-3 px-md-4 mb-3 bg-white border-bottom box-shadow"><h5 class="my-0 mr-md-auto font-weight-normal">'+d+'</h5> </div> <div class="mx-5"> <table class="tablesorter-bootstrap" id="myTable"> <thead> <tr> <th scope="col">日付</th> <th scope="col">学籍番号</th> <th scope="col">クラス</th> <th scope="col">番号</th> <th scope="col">氏名</th> </tr> </thead> <tbody> '
    #Excelを読み込み
    wb = openpyxl.load_workbook('Book1.xlsx')
    #Excelのシート読み込み
    ws = wb[d_now]

    #出席番号順にソート
    temp_number = []
    list_date = []
    list_id = []
    list_class = []
    list_number = []
    list_name = []

    new_date = []
    new_id = []
    new_class = []
    new_number = []
    new_name = []

    for row in ws.iter_rows(min_row=1):
        data_date = row[0].value
        list_date.append(data_date)
        data_id = int(row[1].value)
        list_id.append(data_id)
        data_class = row[2].value
        list_class.append(data_class)
        data_number = int(row[3].value)
        list_number.append(data_number)
        temp_number.append(data_number)
        data_name = row[4].value
        list_name.append(data_name)

    temp_number.sort()
    print(temp_number)

    for row in range(len(temp_number)):
        for i in range(len(list_number)):
            if list_number[i] == temp_number[row]:
                new_date.append(list_date[i])
                new_id.append(list_id[i])
                new_class.append(list_class[i])
                new_number.append(list_number[i])
                new_name.append(list_name[i])
                break
        
    
    #HTMLテーブル作成
    for i in range(len(new_id)):
        datalist += '<tr> <td scope="row"><div class="text-center">'+str(new_date[i])+'</div></td><td><div class="text-center">'+str(new_id[i])+'</div></td><td><div class="text-center">'+str(new_class[i])+'</div></td><td><div class="text-center">'+str(new_number[i])+'</div></td><td><div class="text-center">'+str(new_name[i])+'</div></td></tr>'
    
    datalist += ' </tbody> </table> </div> </body>'
    #HTML書き出し
    f.writelines(datalist)
    #HTML閉じる
    f.close()

    print("html-ok")

#SMBにPDFを送信
def sendPDF(d):
    #ファイルパス
    output_file = 'pdf/'+str(d)+'.pdf'
    #いろいろと設定
    conn = SMBConnection(
        'attendance_check',
        'attendance_check',
        platform.uname().node,
        '<remote_hostname>',
        domain='WORKGROUP',
        use_ntlm_v2=True)
    #SMB接続
    conn.connect('10.65.134.149', 139)

    print(conn.echo('echo success'))

    #pdfを送信
    with open(output_file,'rb') as file:
        print(output_file)
        conn.storeFile('pdf', str(d)+".pdf", file)

    #SMB接続解除
    conn.close()

#SMBにPDFを送信
def sendHTML(d):
    #ファイルパス
    output_file = '_html/'+str(d)+'.html'
    #いろいろと設定
    conn = SMBConnection(
        'attendance_check',
        'attendance_check',
        platform.uname().node,
        '<remote_hostname>',
        domain='WORKGROUP',
        use_ntlm_v2=True)
    #SMB接続
    conn.connect('10.65.134.149', 139)

    print(conn.echo('echo success'))

    #HTMLを送信
    with open(output_file,'rb') as file:
        print(output_file)
        conn.storeFile('html', str(d)+".html", file)

    #SMB接続解除
    conn.close()

#NFC読み込み
def readPaSoRi():
    global id16
    global id10
    tag = clf.connect(rdwr={'on-connect': lambda tag: False})
    id16 = tag.identifier.hex().upper()
    id10 = int(id16,16)

#表に名前を登録する
def name (input_id) :
    #Excelを読み込み
    wb = openpyxl.load_workbook('Book1.xlsx')
    #Excelのシート読み込み
    ws = wb['menber']
    ws2 = wb[d_now]
    for row in ws.iter_rows(min_row=1):
        dt_now = datetime.datetime.now()
        d = str(dt_now.strftime('%Y-%m-%d %H-%M-%S'))
        d2 = str(dt_now.strftime('%Y/%m/%d %H:%M:%S'))
        # 学籍番号
        student_id = row[1].value
        # クラス
        student_class = row[2].value
        # 番号
        student_number = row[3].value
        # 名前
        student_name = row[4].value
        #発見
        if str(student_id) == str(input_id):
            print("日時："+str(d2))
            print("学籍番号："+str(student_id))
            print("クラス："+str(student_class))      
            print(str(student_name)+"　サマ")
            ws2.append([d2,student_id,student_class,student_number,student_name])
            wb.save('Book1.xlsx')
            return student_name

    else:
        return 0 

#表から学生番号を取得
def barcode (input_id) :
    #Excelを読み込み
    wb = openpyxl.load_workbook('Book1.xlsx')
    #Excelのシート読み込み
    ws = wb['menber']
    sheet = wb.worksheets[0]
    max =  ws.max_row
    for row in ws.iter_rows(min_row=2):
        # 学生商Id
        member_id = row[0].value
        # 学籍番号
        student_number = row[1].value

        #発見
        if member_id == input_id:
            print ("学籍番号",student_number)
            return (student_number)
    else:
        return 0    

#バーコードを使ったNFCタグ登録
def register (input_id, input_member) :
    print(input_id)
    print(input_member)
    #Excelを読み込み
    wb = openpyxl.load_workbook('Book1.xlsx')
    #Excelのシート読み込み
    ws1 = wb['menber']
    ws3 = wb['名簿']
    for row in ws3.iter_rows(min_row=1):
        
     # 学籍番号
        data_member = row[0].value

        #発見
        if str(data_member) == str(input_member):
          data_class = row[1].value
          data_number = row[2].value
          data_name = row[3].value
          break
    else:
        print("エラー(データなし)")    
    ws1.append([input_id,input_member,data_class,data_number,data_name])
    wb.save('Book1.xlsx')
    print('complete save')

    return

target = clf.sense(RemoteTarget('106A'))

#メイン関数
if __name__ == "__main__":
    #Qt立ち上げ
    app = QtWidgets.QApplication(sys.argv)
    Form = QtWidgets.QWidget()
    ui = Ui_Form()
    ui.setupUi(Form)
    Form.show()

    #現在時刻
    global d_now
    dt_now = datetime.datetime.now()
    d_now = str(dt_now.strftime('%Y-%m-%d-%H-%M-%S'))
    #Excelブック呼び出し
    wb = openpyxl.load_workbook('Book1.xlsx')
    #現在時刻のシートを作成
    wb.create_sheet(d_now)
    wb.save('Book1.xlsx')

    global id10
    global id16
    #ID初期化
    id16 = int('0xff', 16)
    id10 = 1100000
    old_ID = int('0xff', 16)

    #NFCリーダースレッド立ち上げ
    thread1 = threading.Thread(target=readPaSoRi)
    thread1.start()

    @with_goto
    def update_label():
        global end
        thread1 = threading.Thread(target=readPaSoRi)
        if thread1.isAlive() == False:
            thread1.start()
        global old_ID
        dt_now = datetime.datetime.now()
        d = str(dt_now.strftime('%Y/%m/%d %H:%M:%S'))
        # current_time = str(datetime.datetime.now().time())
        # old_registernumber = int('EEEFFF',16)

        #読み取るNFCIDが重ならないように
        if old_ID != id16:
            #終了専用のNFCタグ設定
            if(id16 == 'B685C23B'):
                #終了音再生
                playsound("Ticket_Gate-Beep01-07(Tone2).mp3")
                dt_now = datetime.datetime.now()
                d3 = str(dt_now.strftime('%Y-%m-%d_%H-%M-%S'))
                print("end")
                #HTML/PDFを作成・送信
                makePDF(d3)
                makeHTML(d3)
                sendPDF(d3)
                sendHTML(d3)
                #システム終了
                sys.exit(app.exec_())
                return
            else:
                return_comment = barcode(id16)
                #NFCタグ登録されているか
                if return_comment == 0:
                    #未踏特
                    playsound("Ticket_Gate-Beep01-03(Tone1).mp3")
                    ui.label.setText("登録が必要です。バーコードを読み取らせてください")
                    ui.label.setFont(QFont('Arial',20))
                    label.begin
                    register_number = designQT.number
                    #入力された数値の桁数によって次のステップへ
                    if len(register_number) == 11:
                        #Qtに描画
                        ui.numberLabel.setText("--------")
                        ui.numberLabel.setFont(QFont('Arial',18))
                        ui.lineEdit.clear()
                        #コンソールにも表示
                        print(register_number)
                        print(len(register_number))
                        #いらない桁の削除
                        register(id16,register_number[3:10])
                        #外部変数numberの初期化
                        designQT.number = 'FFF'
                        time.sleep(0.5)
                        dt_now = datetime.datetime.now()
                        d = str(dt_now.strftime('%Y/%m/%d %H:%M:%S'))
                        #表から学生番号を取得
                        return_comment = (barcode(id16))
                        #名前を取得
                        namae = name(return_comment)
                        #Qt表示
                        ui.label.setText("ようこそ　{}様".format(namae))
                        ui.label.setFont(QFont('Arial',36))
                        ui.dateLabel.setText(d)
                        ui.dateLabel.setFont(QFont('Arial',18))
                        ui.numberLabel.setText(str(return_comment))
                        ui.numberLabel.setFont(QFont('Arial',18))
                else:
                    playsound("Ticket_Gate-Beep01-02(Tone1).mp3")
                    dt_now = datetime.datetime.now()
                    d = str(dt_now.strftime('%Y/%m/%d %H:%M:%S'))
                    #名前を取得
                    namae = name(return_comment)
                    #Qt表示
                    ui.label.setText("ようこそ　{}様".format(namae))
                    ui.label.setFont(QFont('Arial',36))
                    ui.dateLabel.setText(d)
                    ui.dateLabel.setFont(QFont('Arial',18))
                    ui.numberLabel.setText(str(return_comment))
                    ui.numberLabel.setFont(QFont('Arial',18))


            old_ID = id16
        else:
            #苦肉の策のgoto 超危険
            goto.begin
        
    #Qt描画更新
    timer = QTimer()
    timer.timeout.connect(update_label)
    timer.start(100)  # every 100 milliseconds

    sys.exit(app.exec_())